import logging
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from livekit.agents import (
    Agent,
    AgentSession,
    AutoSubscribe,
    JobContext,
    JobProcess,
    WorkerOptions,
    cli,
    metrics,
    RoomInputOptions,
)
from livekit.plugins import (
    cartesia,
    deepgram,
    groq,
    noise_cancellation,
    silero,
)
from livekit.plugins.cartesia import TTS
from livekit.plugins.deepgram import STT
from livekit.plugins.openai import LLM
from livekit.plugins.silero import VAD
from livekit.plugins.turn_detector.multilingual import MultilingualModel

# Load .env
load_dotenv(dotenv_path=".env.local")

# Setup logging
log_path = os.path.expanduser("~/Desktop/voice_agent_debug003.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("voice-agent")

# Excel output
METRICS_XLSX = os.path.expanduser("~/Desktop/metrics_log003.xlsx")

# Setup Excel file
if not os.path.exists(METRICS_XLSX):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    ws.append([
        "Timestamp", "eou_delay", "transcription_delay",
        "ttft", "llm_input_tokens", "llm_output_tokens",
        "tts_ttfb", "tts_duration", "tts_audio_duration",
        "total_latency"
    ])
    wb.save(METRICS_XLSX)

# Store metric fragments until full set is available
metrics_storage = {}

class Assistant(Agent):
    def __init__(self) -> None:
        super().__init__(
            instructions="You are a voice assistant created by LiveKit. Use short, polite, and clear answers.",
            stt=deepgram.STT(),
            llm=groq.LLM(model="llama3-8b-8192"),
            tts=cartesia.TTS(),
            turn_detection=MultilingualModel(),
        )

    async def on_enter(self):
        await self.session.generate_reply(
            instructions="Hi there! How can I help you today?",
            allow_interruptions=True,
        )

# Preload Silero VAD for silence detection
def prewarm(proc: JobProcess):
    try:
        proc.userdata["vad"] = silero.VAD.load()
        logger.info("Silero VAD loaded successfully.")
    except Exception as e:
        logger.error(f" Failed to load Silero VAD: {e}")
        proc.userdata["vad"] = None

# Agent entrypoint
async def entrypoint(ctx: JobContext):
    logger.info(f" Connecting to room: {ctx.room.name}")
    await ctx.connect(auto_subscribe=AutoSubscribe.AUDIO_ONLY)

    participant = await ctx.wait_for_participant()
    logger.info(f" Participant joined: {participant.identity}")

    def on_metrics_collected(event):
        try:
            metric_data = getattr(event, "metrics", event)
            metric_type = getattr(metric_data, "type", type(metric_data).__name__)
            logger.info(f" Received metric type: {metric_type}")
            speech_id = getattr(metric_data, "speech_id", None)
            if not speech_id:
                logger.warning(" Missing speech_id — skipping metric.")
                return

            # Store partial metrics
            if speech_id not in metrics_storage:
                metrics_storage[speech_id] = {}

            row = metrics_storage[speech_id]

            # Capture individual metric values
            if metric_type == "eou_metrics":
                row["eou_delay"] = getattr(metric_data, "end_of_utterance_delay", "")
                row["transcription_delay"] = getattr(metric_data, "transcription_delay", "")
            elif metric_type == "llm_metrics":
                row["ttft"] = getattr(metric_data, "ttft", "")
                row["llm_input_tokens"] = getattr(metric_data, "prompt_tokens", "")
                row["llm_output_tokens"] = getattr(metric_data, "completion_tokens", "")
            elif metric_type == "tts_metrics":
                row["tts_ttfb"] = getattr(metric_data, "ttfb", "")
                row["tts_duration"] = getattr(metric_data, "duration", "")
                row["tts_audio_duration"] = getattr(metric_data, "audio_duration", "")

            # If all important metrics are present, compute total latency and log
            if row.keys() >= {"eou_delay", "transcription_delay", "ttft", "tts_ttfb"}:
                row["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                try:
                    row["total_latency"] = (
                        float(row["eou_delay"]) +
                        float(row["transcription_delay"]) +
                        float(row["ttft"]) +
                        float(row["tts_ttfb"])
                    )
                except Exception as e:
                    logger.warning(f" Could not compute total_latency: {e}")
                    row["total_latency"] = ""

                # Save row to Excel
                wb = load_workbook(METRICS_XLSX)
                ws = wb["Metrics"]
                ws.append([
                    row.get("Timestamp", ""),
                    row.get("eou_delay", ""),
                    row.get("transcription_delay", ""),
                    row.get("ttft", ""),
                    row.get("llm_input_tokens", ""),
                    row.get("llm_output_tokens", ""),
                    row.get("tts_ttfb", ""),
                    row.get("tts_duration", ""),
                    row.get("tts_audio_duration", ""),
                    row.get("total_latency", "")
                ])
                wb.save(METRICS_XLSX)

                logger.info(f" Logged all metrics for speech_id={speech_id}")
                metrics_storage.pop(speech_id, None)

        except Exception as e:
            logger.error(f" Error during metric logging: {e}")

    # Agent session setup
    session = AgentSession(
        vad=ctx.proc.userdata.get("vad", None),
        min_endpointing_delay=0.5,
        max_endpointing_delay=5.0,
    )

    session.on("metrics_collected", on_metrics_collected)

    await session.start(
        room=ctx.room,
        agent=Assistant(),
        room_input_options=RoomInputOptions(
            noise_cancellation=noise_cancellation.BVC(),
        ),
    )

# Start the voice agent
if __name__ == "__main__":
    cli.run_app(
        WorkerOptions(
            entrypoint_fnc=entrypoint,
            prewarm_fnc=prewarm,
        )
    )
